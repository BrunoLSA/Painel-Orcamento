#!/usr/bin/env python3
# =============================================================================
# Importador dos dados reais (Tesouro Gerencial) -> CSVs em data/fonte/
# Converte:
#   1) Planilha "Execução Plano de Ação" (aba unica)
#   2) Planilha "RP" (aba "BASE RESTOS A PAGAR")
# em: execucao.csv, credito_diref.csv, credito_ugr.csv,
#     restos_a_pagar.csv, rap_ugr.csv, execucao_ugr.csv
#
# Uso: python3 scripts/importar_tg.py <execucao.xlsx> <rp.xlsx> [ano] [mes]
#   ano padrao = 2026 ; mes padrao = "JUN/2026" (mes mais recente / acumulado)
#
# Regras (definidas com a area):
#   - Execucao: usa o mes mais recente (valores acumulados).
#   - recebido = CREDITO DISPONIVEL + DESPESAS EMPENHADAS.
#   - cambio   = DOTACAO das linhas com PI = 00000CAMBIO.
#   - Credito DIREF = credito disponivel retido na DIREF (sem UGR), por AO/ND/Fonte.
#   - Credito UGR   = credito disponivel descentralizado, por UGR/AO.
#   - RP: apenas RPNP (Bloco B, col 33-40). saldo a pagar = A Liquidar + Liq. a Pagar.
# =============================================================================
import sys, csv, os
from collections import defaultdict
import openpyxl

ANO = sys.argv[3] if len(sys.argv) > 3 else "2026"
MES = sys.argv[4] if len(sys.argv) > 4 else "JUN/2026"
EXEC_XLSX = sys.argv[1]
RP_XLSX = sys.argv[2]
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "fonte")

def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0

def s(v):
    return ("" if v is None else str(v)).strip()

def retido(cod, nome):
    """True se o credito ainda esta na DIREF (nao descentralizado a uma UGR)."""
    c = cod.replace("'", "").strip()
    return c in ("", "-8") or "SEM INFORMACAO" in nome.upper()

# ----- 1) EXECUCAO (mes mais recente) ---------------------------------------
wb = openpyxl.load_workbook(EXEC_XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
linhas = []
for row in ws.iter_rows(min_row=8, values_only=True):
    if not row or s(row[0]) == "":
        continue
    if s(row[1]) != MES:
        continue
    disp, emp = num(row[14]), num(row[15])
    linhas.append({
        "ao": s(row[0]),
        "ugr_cod": s(row[3]).replace("'", ""),
        "ugr_nome": s(row[4]),
        "pi": s(row[6]),
        "fonte": s(row[8]), "fonte_nome": s(row[9]),
        "nd": s(row[10]), "nd_nome": s(row[11]),
        "dotacao": num(row[13]),
        "disponivel": disp,
        "empenhado": emp,
        "liquidado": num(row[16]),
        "pago": num(row[17]),
        "recebido": disp + emp,
    })
wb.close()

# Nomes de AO a partir da planilha de RP
wb = openpyxl.load_workbook(RP_XLSX, read_only=True, data_only=True)
ws = wb["BASE RESTOS A PAGAR"]
ao_nome = {}
rp = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row:
        continue
    ao = s(row[11]) or s(row[3])
    if ao == "":
        continue
    nome = s(row[12])
    if ao and nome and ao not in ao_nome:
        ao_nome[ao] = nome
    rp.append({
        "ao": ao, "ao_nome": nome,
        "ugr_cod": s(row[8]).replace("'", ""), "ugr_nome": s(row[9]),
        "inscrito": num(row[32]),   # col33 INSCRITOS+REINSCRITOS (bloco B)
        "cancelado": num(row[35]),  # col36 CANCELADOS
        "a_liquidar": num(row[36]), # col37 A LIQUIDAR
        "liq_a_pagar": num(row[37]),# col38 LIQUIDADOS A PAGAR
        "pago": num(row[38]),       # col39 PAGOS
    })
wb.close()

def nome_ao(ao):
    return ao_nome.get(ao, ao)

os.makedirs(OUT, exist_ok=True)
def escrever(nome, cab, linhas_csv):
    with open(os.path.join(OUT, nome), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(cab)
        w.writerows(linhas_csv)
    print(f"  {nome}: {len(linhas_csv)} linhas")

# ----- execucao.csv (por AO) -------------------------------------------------
ex = defaultdict(lambda: dict(dotacao=0, recebido=0, empenhado=0, liquidado=0, pago=0, cambio=0))
for l in linhas:
    a = ex[l["ao"]]
    for k in ("dotacao", "recebido", "empenhado", "liquidado", "pago"):
        a[k] += l[k]
    if l["pi"].upper() == "00000CAMBIO":
        a["cambio"] += l["dotacao"]
rows = [[ANO, ao, nome_ao(ao), round(v["dotacao"], 2), round(v["recebido"], 2),
         round(v["empenhado"], 2), round(v["liquidado"], 2), round(v["pago"], 2), round(v["cambio"], 2)]
        for ao, v in sorted(ex.items(), key=lambda x: -x[1]["dotacao"])]
escrever("execucao.csv", ["ano","acao","acao_nome","dotacao","recebido","empenhado","liquidado","pago","cambio"], rows)

# ----- credito_diref.csv (credito disponivel detalhado por AO/ND/Fonte) ------
df = defaultdict(float)
for l in linhas:
    if l["disponivel"] != 0:
        df[(l["ao"], l["nd"], l["nd_nome"], l["fonte"])] += l["disponivel"]
rows = [[ANO, ao, nd, ndn, fonte, "", round(v, 2)]
        for (ao, nd, ndn, fonte), v in sorted(df.items(), key=lambda x: -x[1])]
escrever("credito_diref.csv", ["ano","acao","nd","nd_nome","fonte","ptres","disponivel"], rows)

# ----- credito_ugr.csv (credito disponivel descentralizado, por UGR/AO) ------
cu = defaultdict(float)
unome = {}
for l in linhas:
    if not retido(l["ugr_cod"], l["ugr_nome"]) and l["disponivel"] != 0:
        cu[(l["ugr_cod"], l["ao"])] += l["disponivel"]
        unome[l["ugr_cod"]] = l["ugr_nome"]
rows = [[ANO, cod, cod, unome.get(cod, cod), ao, round(v, 2)]
        for (cod, ao), v in sorted(cu.items(), key=lambda x: -x[1])]
escrever("credito_ugr.csv", ["ano","ugr_codigo","ugr_sigla","ugr_nome","acao","disponivel"], rows)

# ----- execucao_ugr.csv (execucao por UGR; apenas unidades descentralizadas) -
eu = defaultdict(lambda: dict(recebido=0, empenhado=0, liquidado=0, pago=0, nome=""))
for l in linhas:
    if retido(l["ugr_cod"], l["ugr_nome"]):
        continue
    g = eu[l["ugr_cod"]]; g["nome"] = l["ugr_nome"]
    for k in ("recebido", "empenhado", "liquidado", "pago"):
        g[k] += l[k]
rows = [[ANO, cod, cod, v["nome"], round(v["recebido"],2), round(v["empenhado"],2), round(v["liquidado"],2), round(v["pago"],2)]
        for cod, v in sorted(eu.items(), key=lambda x: -x[1]["recebido"])]
escrever("execucao_ugr.csv", ["ano","ugr_codigo","ugr_sigla","ugr_nome","recebido","empenhado","liquidado","pago"], rows)

# ----- restos_a_pagar.csv (apenas RPNP) --------------------------------------
tot = dict(inscrito=0, cancelado=0, a_liquidar=0, liq_a_pagar=0, pago=0)
for r in rp:
    for k in tot:
        tot[k] += r[k]
liquidado = tot["liq_a_pagar"] + tot["pago"]
escrever("restos_a_pagar.csv", ["ano","tipo","sigla","inscrito","cancelado","liquidado","pago"],
         [[ANO, "Não Processados", "RPNP", round(tot["inscrito"],2), round(tot["cancelado"],2),
           round(liquidado,2), round(tot["pago"],2)]])

# ----- rap_ugr.csv (saldo a pagar por UGR/AO) --------------------------------
ru = defaultdict(float)
rnome = {}
for r in rp:
    cod = r["ugr_cod"] or "-"
    saldo = r["a_liquidar"] + r["liq_a_pagar"]
    if saldo == 0:
        continue
    ru[(cod, r["ao"])] += saldo
    rnome[cod] = r["ugr_nome"] or cod
rows = [[ANO, cod, cod, rnome.get(cod, cod), ao, round(v, 2)]
        for (cod, ao), v in sorted(ru.items(), key=lambda x: -x[1])]
escrever("rap_ugr.csv", ["ano","ugr_codigo","ugr_sigla","ugr_nome","acao","a_pagar"], rows)

print(f"\nExercicio={ANO}  Mes(execucao)={MES}")
print(f"Execucao: dotacao total = {sum(l['dotacao'] for l in linhas):,.2f}")
print(f"RP RPNP inscrito = {tot['inscrito']:,.2f} | saldo a pagar = {tot['a_liquidar']+tot['liq_a_pagar']:,.2f}")
